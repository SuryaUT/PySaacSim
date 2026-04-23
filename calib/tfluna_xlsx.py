"""Fit per-lidar TFLuna calibration from a sheet of (true distance, reported
distance per lidar) measurements.

Two xlsx layouts are supported:

  * Modern `Sensor_Calib.xlsx`, sheet `LiDARs`:
      inches, mm, Left LiDAR, Center LiDAR, Right LiDAR

  * Legacy / synthetic:
      true_mm, tf_left, tf_middle, tf_right  (or with `_mm` suffixes)

True distance preference: a `mm` column wins over `true_mm` wins over
`inches * 25.4`. Per-lidar columns are matched by case-insensitive
substring on the **lidar identifier** (`left`, `center` / `middle`,
`right`); the tokenization avoids matching IR sheets by requiring
either the column header to also contain `lidar`/`tf`, OR the sheet
name itself to contain `lidar`/`tf`.

Output keys are firmware-faithful: `tf_left`, `tf_middle`, `tf_right`
(matches `sim/sensors.py` placement IDs `tfluna_left`,
`tfluna_center`, `tfluna_right` minus the `tfluna_` prefix).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

import numpy as np

import openpyxl


@dataclass
class TFLunaSideFit:
    id: str
    scale: float
    bias_cm: float
    noise_std_cm: float
    n_points: int


@dataclass
class TFLunaXlsxFit:
    per_lidar: dict[str, TFLunaSideFit]


def _find_tfluna_sheet(wb) -> Optional[tuple[str, list[tuple]]]:
    """Pick a sheet that looks like TFLuna calibration. Two acceptance paths:
       1. The sheet name itself contains "lidar" or "tf".
       2. At least one header cell contains "lidar" or "tf"/"tfluna".
    Either path requires the data rows to have a usable "true distance"
    column ("mm" or "true" substring) — otherwise we can't fit.
    """
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] is not None]
        if not rows:
            continue
        sheet_name_lc = name.strip().lower()
        header = [str(c).strip().lower() for c in rows[0] if c is not None]
        sheet_match = "lidar" in sheet_name_lc or "tf" in sheet_name_lc
        header_match = any(("tf" in h or "tfluna" in h or "lidar" in h) for h in header)
        if not (sheet_match or header_match):
            continue
        return name, rows
    return None


def _is_tfluna_header(h: str) -> bool:
    """Header context check — is this a column that's plausibly per-lidar
    rather than something else (IR ADC, throttle, etc.)? True iff it
    contains 'lidar' or 'tf'."""
    return ("lidar" in h) or ("tf" in h)


def _opt_col(header: list[str], substr: str) -> Optional[int]:
    for i, h in enumerate(header):
        if substr in h:
            return i
    return None


def fit_xlsx(path: Union[str, Path]) -> Optional[TFLunaXlsxFit]:
    p = Path(path)
    wb = openpyxl.load_workbook(p, data_only=True)
    found = _find_tfluna_sheet(wb)
    if found is None:
        return None
    _, rows = found
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    data = rows[1:]

    # Resolve true-distance column (mm). Preference order:
    #   1. column with 'mm' but not 'lidar' (canonical Sensor_Calib layout)
    #   2. column with 'true' (legacy)
    #   3. column with 'inch' → convert with *25.4
    i_true_mm = None
    for i, h in enumerate(header):
        if "mm" in h and not _is_tfluna_header(h):
            i_true_mm = i
            break
    i_true_legacy = _opt_col(header, "true") if i_true_mm is None else None
    i_inch = _opt_col(header, "inch") if (i_true_mm is None and i_true_legacy is None) else None
    if i_true_mm is not None:
        true_mm = np.array([float(r[i_true_mm]) for r in data if r[i_true_mm] is not None])
    elif i_true_legacy is not None:
        true_mm = np.array([float(r[i_true_legacy]) for r in data if r[i_true_legacy] is not None])
    elif i_inch is not None:
        true_mm = np.array([float(r[i_inch]) * 25.4 for r in data if r[i_inch] is not None])
    else:
        return None

    # Per-lidar columns: try modern names first ('left lidar' etc.), fall
    # back to legacy ('tf_left' etc.). Both kinds map to firmware-faithful
    # output keys: tf_left / tf_middle / tf_right.
    label_map = (
        ("tf_left",   ("left lidar",   "tf_left",   "tfluna_left"  )),
        ("tf_middle", ("center lidar", "middle lidar", "center", "tf_middle", "tfluna_middle", "tfluna_center")),
        ("tf_right",  ("right lidar",  "tf_right",  "tfluna_right" )),
    )
    fits: dict[str, TFLunaSideFit] = {}
    for out_key, candidates in label_map:
        col_idx: Optional[int] = None
        for cand in candidates:
            idx = _opt_col(header, cand)
            if idx is not None:
                col_idx = idx
                break
        if col_idx is None:
            continue
        reported = np.array(
            [float(r[col_idx]) if r[col_idx] is not None else np.nan for r in data]
        )
        # Drop rows where either true_mm or reported is NaN (length must match).
        good = ~np.isnan(reported)
        if good.sum() < 3:
            continue
        true_clean = true_mm[: reported.size][good]
        rep_clean = reported[good]
        # Linear LS fit: reported = scale * true + bias (mm units).
        A = np.column_stack([true_clean, np.ones_like(true_clean)])
        (scale, bias), *_ = np.linalg.lstsq(A, rep_clean, rcond=None)
        resid = rep_clean - (scale * true_clean + bias)
        fits[out_key] = TFLunaSideFit(
            id=out_key,
            scale=float(scale),
            bias_cm=float(bias) / 10.0,
            noise_std_cm=float(np.std(resid)) / 10.0,
            n_points=int(rep_clean.size),
        )
    if not fits:
        return None
    return TFLunaXlsxFit(per_lidar=fits)
